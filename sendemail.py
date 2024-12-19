import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate
from dotenv import load_dotenv
import os
import openpyxl
from datetime import datetime
import re
import random

# Load environment variables from .env file
load_dotenv("credentials.env")

# SMTP server details from .env file
smtp_host = os.getenv('SMTP_HOST')
smtp_port = int(os.getenv('SMTP_PORT', 25))  # Use default port 25 if not found
sender_email = os.getenv('SENDER_EMAIL')
sender_password = os.getenv('SENDER_PASSWORD')
bcc_email = os.getenv('BCC_EMAIL')

# Debugging: Verify credentials loaded correctly
print(f"SMTP Host: {smtp_host}, Port: {smtp_port}, Sender Email: {sender_email}")

# Open the Excel file containing email details
workbook = openpyxl.load_workbook('emails.xlsx', data_only=True)
sheet = workbook['Sheet1']

# Create a new workbook for processed emails
processed_workbook = openpyxl.Workbook()
processed_sheet = processed_workbook.active
processed_sheet.title = "Processed Emails"

# Add headers to the processed emails sheet
processed_sheet.append(["Email", "Company", "Status", "Timestamp"])

# Function to check the validity of an email address
def is_valid_email(email):
    if email:
        email = email.strip()  # Remove leading/trailing spaces
        return re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email) is not None
    return False

# Function to log email status in the processed emails workbook
def log_email_status(email, company, status):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    processed_sheet.append([email, company, status, timestamp])

# Extract recipient details from the Excel file
recipient_details = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
    company = row[0]  # Assuming Company Name is in Column A
    email = row[2]    # Assuming Email Address is in Column C
    if is_valid_email(email):
        recipient_details.append((email, company))
    else:
        print(f"Invalid email: {email} for company {company}")

# Print all the emails that will be sent out at the start
print("\nEmails to be sent out:")
for email, company in recipient_details:
    print(f"- {email} (Company: {company})")

# Connect to the SMTP server
try:
    server = smtplib.SMTP(smtp_host, smtp_port)
    server.starttls()
    server.login(sender_email, sender_password)
    print("\nSMTP server connected successfully.\n")
except Exception as e:
    print(f"Failed to connect to SMTP server: {e}")
    exit()

# Open a status log file with UTF-8 encoding for writing
with open('status.txt', 'a', encoding='utf-8') as status_file:
    # Loop through recipient details and send emails
    for index, (email, company) in enumerate(recipient_details, start=2):  # Start from row 2
        try:
            # Create the email message
            message = MIMEMultipart()
            
            # Set the sender with a display name
            sender_display_name = "Ben"
            message['From'] = formataddr((sender_display_name, sender_email))            
            message['To'] = email
            message['Bcc'] = bcc_email
            subject_company = company if company else "Your Industry"
            message['Subject'] = f'One-Stop Solution for High-Quality Panels and Enclosures for {subject_company}'
            message['Date'] = formatdate(localtime=True)

            # Read the email body content from a text file
            with open('message.txt', 'r', encoding='utf-8') as file:
                body = file.read()
            message.attach(MIMEText(body, 'plain', 'utf-8'))  # Ensure UTF-8 encoding

            # Attach the product brochure
            file_path = 'product_brochure.pdf'
            with open(file_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{file_path}"')
            message.attach(part)

            # Send the email
            server.sendmail(sender_email, [email, bcc_email], message.as_string())
            print(f"Email sent to {email}")

            # Log success in both original and processed files
            log_email_status(email, company, "Sent")
            status_file.write(f"Email sent to {email}\n")

        except UnicodeEncodeError as ue:
            print(f"Encoding error sending email to {email}: {ue}")
            log_email_status(email, company, f"Encoding error: {ue}")
            status_file.write(f"Encoding error sending email to {email}: {ue}\n")
            continue  # Move to the next email
        except Exception as e:
            print(f"Error sending email to {email}: {e}")
            log_email_status(email, company, f"Error: {e}")
            status_file.write(f"Error sending email to {email}: {e}\n")
            continue  # Move to the next email

        # Add a random delay between emails to avoid spamming
        wait_time = random.randint(150, 200)  # Delay between 1 to 2 minutes before sending
        print(f"Waiting for {wait_time}s")
        time.sleep(wait_time)

# Save the updated Excel files
workbook.save('emails_sent.xlsx')
processed_workbook.save('processed_emails.xlsx')

# Disconnect from the SMTP server
server.quit()
print("\nAll emails processed and server connection closed.")
